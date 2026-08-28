import streamlit as st
import pdfplumber, openpyxl, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

# ─── Constants ────────────────────────────────────────────────────────────────
SIZE_TOKENS = (
    'K4','K6','K8','K10','K12','K14','K16',
    'XXS','XS','S','M','L','XL',
    '2XL','3XL','4XL','5XL','6XL','7XL','8XL',
    'L6','L8','L10','L12','L14','L16','L18','L20','L22','L24','L26',
    'TO 0','TO 1','TO 2','TO 3','TO 4','TO 5','TO 6',
    'One'
)
SIZE_TOKENS_SET = set(SIZE_TOKENS)

LINE_COLORS = [
    'FF6D01','FFFF00','FF6D01','F6B3AE',
    'C2EF51','FFE599','ADD8E6','D8BFD8','90EE90','FFDAB9'
]

# ─── Core parsing helpers (unchanged logic) ───────────────────────────────────
def get_order_no(words):
    for i, w in enumerate(words):
        if w['text'] == 'Order' and i+1 < len(words) and words[i+1]['text'] in ('No', 'Number'):
            for j in range(i+2, min(i+8, len(words))):
                if words[j]['text'] in (':', ' ', 'No', 'Number'):
                    continue
                cleaned = words[j]['text'].replace(',', '')
                if len(cleaned) >= 4 and cleaned.isdigit():
                    return cleaned
    return ''

def get_line_no(words):
    for i, w in enumerate(words):
        if w['text'] == 'Line' and i+1 < len(words) and words[i+1]['text'] in ('No', 'Number'):
            for j in range(i+2, min(i+8, len(words))):
                val = words[j]['text'].replace(':', '').strip()
                if val.isdigit() and len(val) < 4:
                    return val
    return None

def merge_size_tokens(words):
    result = []
    i = 0
    while i < len(words):
        w = words[i]
        if w['text'].upper() == 'TO' and i + 1 < len(words):
            combined = f"TO {words[i+1]['text']}"
            if combined in SIZE_TOKENS_SET:
                merged = dict(w); merged['text'] = combined
                result.append(merged); i += 2; continue
        result.append(w); i += 1
    return result

def is_name_number_page(wbt, words):
    if not all(k in wbt for k in ('Name','Number','Qty','Size')):
        return False
    rows = defaultdict(list)
    for key in ('Name','Number','Qty','Size'):
        for w in wbt[key]:
            rows[round(w['top']/4)*4].append(w['text'])
    return any(all(t in rt for t in ('Name','Number','Qty','Size')) for rt in rows.values())

def is_names_grid_page(wbt):
    return all(k in wbt for k in ('Names','Sizes','Qunatities'))

def is_name_requirements_page(wbt):
    return all(k in wbt for k in ('Name','Requirements','Number','Size'))

def parse_names_grid_page(words, wbt):
    size_header_rows = {}
    for w in words:
        if w['text'] in SIZE_TOKENS_SET:
            size_header_rows.setdefault(round(w['top']/4)*4, []).append(w)
    if not size_header_rows:
        return []
    sorted_tops = sorted(size_header_rows)
    inner_top = next((ht for ht in sorted_tops if ht > 130), sorted_tops[-1])
    header_sizes = sorted(size_header_rows[inner_top], key=lambda w: w['x0'])
    total_words = [w for w in wbt.get('Total',[]) if abs(round(w['top']/4)*4 - inner_top) < 8]
    total_x = total_words[0]['x0'] if total_words else float('inf')
    excluded = {'Names','Sizes','Qunatities','Total','Printed','Page','of',
                ':','Type','Order','Details','Customer','ID','Ship','By',
                'Date','No','Line','AJAX','JFC'}
    name_words = [w for w in words
                  if w['x0'] < 80 and w['top'] > inner_top + 10
                  and w['text'] not in excluded
                  and not w['text'].replace(':','').strip().isdigit()
                  and w['text'] not in SIZE_TOKENS_SET
                  and len(w['text']) > 1
                  and not w['text'].startswith('2')
                  and ':' not in w['text']]
    name_rows = defaultdict(list)
    for w in name_words:
        name_rows[round(w['top']/4)*4].append(w)
    digit_by_row = defaultdict(list)
    for w in words:
        if w['text'].isdigit() and w['x0'] < total_x - 5:
            digit_by_row[round(w['top']/4)*4].append(w)
    entries = []
    for rk, rnw in name_rows.items():
        rnw_s = sorted(rnw, key=lambda w: w['x0'])
        name = ' '.join(w['text'] for w in rnw_s)
        name_top = rnw_s[0]['top']
        for dw in digit_by_row[round(name_top/4)*4]:
            nearest = min(header_sizes, key=lambda s: abs(s['x0']-dw['x0']))
            if abs(nearest['x0']-dw['x0']) < 40:
                try:
                    entries.append({'size':nearest['text'],'name':name,'number':None,'qty':int(dw['text'])})
                except ValueError:
                    pass
    return entries

def parse_name_number_page(words, wbt):
    header_words = [w for k in ('Size','Name','Number','Qty') for w in wbt.get(k,[])]
    if not header_words:
        return []
    header_top = min(w['top'] for w in header_words)
    left_h  = sorted([w for w in header_words if w['x0'] <  350 and abs(w['top']-header_top)<8], key=lambda w: w['x0'])
    right_h = sorted([w for w in header_words if w['x0'] >= 350 and abs(w['top']-header_top)<8], key=lambda w: w['x0'])
    skip = {'Printed',':','Page','of','Total'}

    def parse_group(headers, all_words, min_x, max_x):
        if len(headers) < 4: return []
        col_x = {h['text']: h['x0'] for h in headers}
        col_coords = sorted([(k, col_x[k]) for k in ('Size','Name','Number','Qty') if k in col_x], key=lambda x: x[1])
        data = [w for w in all_words if w['top']>header_top+5 and min_x<=w['x0']<=max_x and w['text'] not in skip]
        prox_rows = []
        for w in sorted(data, key=lambda w: w['top']):
            placed = False
            for pr in prox_rows:
                if abs(w['top'] - pr[0]['top']) <= 10:
                    pr.append(w); placed = True; break
            if not placed:
                prox_rows.append([w])
        entries = []
        for prox_row in prox_rows:
            rw = sorted(prox_row, key=lambda w: w['x0'])
            crd = defaultdict(list)
            for w in rw:
                col = min(col_coords, key=lambda c: abs(c[1]-w['x0']))
                if abs(col[1]-w['x0']) < 60:
                    crd[col[0]].append(w['text'])
            rd = {k: ' '.join(crd[k]).strip() if crd[k] else None for k in ('Size','Name','Number','Qty')}
            if rd['Size'] and rd['Size'] in SIZE_TOKENS_SET:
                try:
                    qty = int(rd['Qty']) if rd['Qty'] and rd['Qty'].isdigit() else 1
                    num = int(rd['Number']) if rd['Number'] and rd['Number'].isdigit() else None
                    entries.append({'size':rd['Size'],'name':rd['Name'],'number':num,'qty':qty})
                except ValueError:
                    pass
        return entries

    return parse_group(left_h, words, 0, 370) + parse_group(right_h, words, 370, 900)

def parse_name_requirements_page(words, wbt):
    header_words = [w for k in ('Number','Size','Name') for w in wbt.get(k,[])]
    if not header_words: return []
    rbt = defaultdict(list)
    for w in header_words: rbt[round(w['top']/4)*4].append(w)
    header_top = next((tk for tk, ws in rbt.items()
                       if all(t in [w['text'] for w in ws] for t in ('Number','Size','Name'))), None)
    if header_top is None: return []
    col_ws = [w for w in header_words if abs(round(w['top']/4)*4-header_top)<8]
    col_x  = {w['text']:w['x0'] for w in col_ws}
    num_x, size_x, name_x = col_x.get('Number',36), col_x.get('Size',90), col_x.get('Name',147)
    skip = {'Number','Size','Name','Requirements','Printed','Page','of','Total',':',
            'Customer','JAMBEROO','SUPEROOS','RLFC'}
    data = [w for w in words if w['top'] > header_top/1 + 8 and w['text'] not in skip]
    data_rows = defaultdict(list)
    for w in data: data_rows[round(w['top']/4)*4].append(w)
    entries = []
    for tk in sorted(data_rows):
        rw = sorted(data_rows[tk], key=lambda w: w['x0'])
        if any(t in [w['text'] for w in rw] for t in ('Printed','Apr','Page')): continue
        merged = []
        skip_next = False
        for idx, w in enumerate(rw):
            if skip_next: skip_next=False; continue
            if w['text']=='TO' and idx+1<len(rw):
                ct = f"TO {rw[idx+1]['text']}"
                if ct in SIZE_TOKENS_SET:
                    mw=dict(w); mw['text']=ct; merged.append(mw); skip_next=True; continue
            merged.append(w)
        num_val=size_val=None; name_parts=[]
        for w in merged:
            dists = [abs(w['x0']-num_x), abs(w['x0']-size_x), abs(w['x0']-name_x)]
            md = min(dists)
            if md > 80: continue
            if dists[0]==md: num_val=w['text']
            elif dists[1]==md: size_val=w['text']
            else: name_parts.append(w['text'])
        if not size_val or not num_val or size_val not in SIZE_TOKENS_SET: continue
        number = None if num_val in ('N/A','n/a','') else (int(num_val) if num_val.isdigit() else None)
        entries.append({'size':size_val,'name':' '.join(name_parts).strip() or None,'number':number,'qty':1})
    return entries

def process_page(page_num, page, total_pages):
    words_raw = page.extract_words(x_tolerance=3, y_tolerance=3, keep_blank_chars=False, use_text_flow=False)
    if len(words_raw) < 5:
        return page_num, None
    words = merge_size_tokens(words_raw)
    wbt = defaultdict(list)
    for w in words: wbt[w['text']].append(w)
    return page_num, {'order_no': get_order_no(words), 'line_no': get_line_no(words), 'words': words, 'wbt': wbt}

def parse_po(pdf_bytes, progress_cb=None):
    order_no = ''
    pages_data = {}
    last_known_line_no = None

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        total = len(pdf.pages)
        page_results = {}
        with ThreadPoolExecutor(max_workers=4) as ex:
            fmap = {ex.submit(process_page, n, p, total): n for n, p in enumerate(pdf.pages, 1)}
            done = 0
            for fut in as_completed(fmap):
                done += 1
                if progress_cb:
                    progress_cb(done, total)
                pn, res = fut.result()
                if res: page_results[pn] = res

        for pn in sorted(page_results):
            r = page_results[pn]
            words, wbt = r['words'], r['wbt']
            ol = r['order_no']
            if ol and len(ol) > len(order_no): order_no = ol
            line_no = r['line_no']
            if not line_no:
                if last_known_line_no: line_no = last_known_line_no
                else: continue
            else: last_known_line_no = line_no
            if line_no not in pages_data:
                pages_data[line_no] = {'has_roster':False,'sizes':[],'roster':[],'size_order':[],'type':'grid'}

            if is_name_number_page(wbt, words):
                pages_data[line_no]['type'] = 'name_table'
                pages_data[line_no]['roster'].extend(parse_name_number_page(words, wbt))
                pages_data[line_no]['has_roster'] = True
            elif is_name_requirements_page(wbt):
                pages_data[line_no]['type'] = 'name_requirements'
                pages_data[line_no]['roster'].extend(parse_name_requirements_page(words, wbt))
                pages_data[line_no]['has_roster'] = True
                if 'Sizes' in wbt:
                    _sw = [w for w in words if w['text'] in SIZE_TOKENS_SET]
                    if _sw:
                        _ht = min(w['top'] for w in _sw)
                        _hs = sorted([w for w in _sw if abs(w['top']-_ht)<6], key=lambda w: w['x0'])
                        _tc = [w for w in wbt.get('Total',[]) if abs(w['top']-_ht)<6]
                        _tx = _tc[0]['x0'] if _tc else float('inf')
                        _nw = [w for w in words if w['text'].isdigit() and int(w['text'])>0 and _ht+5<w['top']<_ht+60]
                        _rt = defaultdict(list)
                        for w in _nw: _rt[round(w['top']/4)*4].append(w)
                        _st = sorted(_rt)
                        if _st:
                            _sq = [min(_hs,key=lambda s:abs(s['x0']-nw['x0']))['text']
                                   for nw in _rt[_st[0]] if abs(nw['x0']-_tx)>=20
                                   and abs(min(_hs,key=lambda s:abs(s['x0']-nw['x0']))['x0']-nw['x0'])<40]
                            if _sq:
                                pages_data[line_no]['sizes'] = _sq
                                pages_data[line_no]['size_order'] = [s['text'] for s in _hs]
            elif is_names_grid_page(wbt):
                pages_data[line_no]['type'] = 'names_grid'
                pages_data[line_no]['roster'].extend(parse_names_grid_page(words, wbt))
                pages_data[line_no]['has_roster'] = True
            else:
                size_words = [w for w in words if w['text'] in SIZE_TOKENS_SET]
                if not size_words: continue
                has_sizes = 'Sizes' in wbt
                has_number = 'Number' in wbt
                if has_sizes:
                    ht = min(w['top'] for w in size_words)
                    hs = sorted([w for w in size_words if abs(w['top']-ht)<6], key=lambda w: w['x0'])
                    so = [s['text'] for s in hs]
                    nw = [w for w in words if w['text'].isdigit() and int(w['text'])>0 and ht+5<w['top']<ht+60]
                    rbt2 = defaultdict(list)
                    for w in nw: rbt2[round(w['top']/4)*4].append(w)
                    st = sorted(rbt2)
                    if not st: continue
                    tcw = [w for w in wbt.get('Total',[]) if abs(w['top']-ht)<6]
                    tx = tcw[0]['x0'] if tcw else float('inf')
                    sq = []
                    for nw2 in rbt2[st[0]]:
                        if abs(nw2['x0']-tx)<20: continue
                        near = min(hs, key=lambda s: abs(s['x0']-nw2['x0']))
                        if abs(near['x0']-nw2['x0'])<40: sq.append(near['text'])
                    pages_data[line_no]['sizes'] = sq
                    pages_data[line_no]['has_roster'] = has_number
                    pages_data[line_no]['size_order'] = so
                    if 'Numbering' in wbt:
                        _nl = wbt['Numbering']; _nt = min(w['top'] for w in _nl)
                        _nbs = [w for w in words if w['text'] in SIZE_TOKENS_SET and _nt-5<w['top']<_nt+20]
                        if _nbs:
                            _nbht = min(w['top'] for w in _nbs)
                            _nbhs = sorted([w for w in _nbs if abs(w['top']-_nbht)<6], key=lambda w: w['x0'])
                            _nbtc = [w for w in wbt.get('Total',[]) if abs(w['top']-_nbht)<6]
                            _nbtx = _nbtc[0]['x0'] if _nbtc else float('inf')
                            _nbfx = min(s['x0'] for s in _nbhs)
                            _nbtops = [w['top'] for w in wbt.get('Total',[]) if w['top']>_nbht+5]
                            _nbad = [w for w in words if w['text'].isdigit() and w['top']>_nbht+5]
                            def _on_tot(top): return any(abs(top-tt)<8 for tt in _nbtops)
                            _nbj  = [w for w in _nbad if w['x0']<_nbfx-5 and not _on_tot(w['top'])]
                            _nbq  = [w for w in _nbad if w['x0']>=_nbfx-5 and not _on_tot(w['top']) and abs(w['x0']-_nbtx)>15]
                            for _jw in _nbj:
                                _num = int(_jw['text'])
                                for _qw in _nbq:
                                    if abs(_qw['top'] - _jw['top']) <= 10:
                                        _nr = min(_nbhs, key=lambda s: abs(s['x0']-_qw['x0']))
                                        if abs(_nr['x0']-_qw['x0'])<40 and abs(_qw['x0']-_nbtx)>15:
                                            pages_data[line_no]['roster'].append({'size':_nr['text'],'name':None,'number':_num,'qty':int(_qw['text'])})
                            if not pages_data[line_no]['sizes'] and _nbhs:
                                pages_data[line_no]['sizes'] = [s['text'] for s in _nbhs]
                else:
                    ht  = min(w['top'] for w in size_words)
                    hs  = sorted([w for w in size_words if abs(w['top']-ht)<6], key=lambda w: w['x0'])
                    fsx = min(s['x0'] for s in hs)
                    total_tops = [w['top'] for w in wbt.get('Total',[]) if w['top']>ht]
                    tcw2= [w for w in wbt.get('Total',[]) if abs(w['top']-ht)<6]
                    tcx = tcw2[0]['x0'] if tcw2 else float('inf')
                    nw2 = [w for w in words if w['text'].isdigit() and w['top']>ht+5]
                    # Exclude words on Total rows (use proximity, not bucket)
                    def on_total_row(top):
                        return any(abs(top - tt) < 8 for tt in total_tops)
                    jn  = [w for w in nw2 if w['x0']<fsx-5 and not on_total_row(w['top'])]
                    qc  = [w for w in nw2 if w['x0']>=fsx-5 and not on_total_row(w['top']) and abs(w['x0']-tcx)>15]
                    # Use proximity row matching (tolerance=10px) instead of fixed bucket
                    for jw in jn:
                        num = int(jw['text'])
                        for qw in qc:
                            if abs(qw['top'] - jw['top']) <= 10:
                                nr = min(hs, key=lambda s: abs(s['x0']-qw['x0']))
                                if abs(nr['x0']-qw['x0'])<40 and abs(qw['x0']-tcx)>15:
                                    try: pages_data[line_no]['roster'].append({'size':nr['text'],'name':None,'number':num,'qty':int(qw['text'])})
                                    except ValueError: pass
                    if not pages_data[line_no].get('size_order'):
                        pages_data[line_no]['size_order'] = [s['text'] for s in hs]

    rows = []
    for ln in sorted(pages_data, key=int):
        d = pages_data[ln]
        label = f'{order_no} line {ln}'
        so = d.get('size_order',[])
        all_s = set(d['sizes']) | {r['size'] for r in d['roster']}
        ord_s = [s for s in so if s in all_s]
        for s in all_s:
            if s not in ord_s: ord_s.append(s)
        by_sz = defaultdict(list)
        for r in d['roster']: by_sz[r['size']].append(r)
        for sz in ord_s:
            rows.append({'line':ln,'label':label,'size':sz,'number':None,'name':None,'scope':'BASE DESIGN'})
            for r in by_sz.get(sz,[]):
                rows.append({'line':ln,'label':label,'size':sz,'number':r['number'],'name':r.get('name'),'scope':'ROSTER','qty':r.get('qty',1)})
    return order_no, rows

def build_excel(order_no_master, rows):
    unique_ids = []
    seen = set()
    for item in rows:
        uid = item.get('subf_number') or item['label'].split(' ')[0]
        if uid not in seen: unique_ids.append(uid); seen.add(uid)
    color_map = {uid: LINE_COLORS[i % len(LINE_COLORS)] for i, uid in enumerate(unique_ids)}
    rql = defaultdict(int)
    for item in rows:
        if item['scope']=='ROSTER': rql[item['label']] += item.get('qty',1)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Sheet1'
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = PatternFill('solid', start_color='002060')
    hfont = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    for col, h in enumerate(['SUBF NUMBER','LINE #','SIZE','NUMBERS','NAMES','SCOPE','VIS NAME'], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font=hfont; c.fill=hf; c.alignment=Alignment(horizontal='center',vertical='center'); c.border=border
    for ri, item in enumerate(rows, 2):
        uid = item.get('subf_number') or item['label'].split(' ')[0]
        rf = PatternFill('solid', start_color=color_map[uid])
        font = Font(name='Arial', size=11)
        ncf = PatternFill('solid', start_color='FFFF00') if item['scope']=='ROSTER' and rql[item['label']]>33 else None
        for col, val in enumerate([item.get('subf_number',''),item['label'],item['size'],item['number'],item['name'],item['scope'],item.get('vis_name','')], 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font=font
            if col<=2: c.fill=rf
            if col in (4,5) and ncf: c.fill=ncf
            c.alignment=Alignment(horizontal='center',vertical='center'); c.border=border
    for col, w in zip('ABCDEFG', [15,22,10,12,18,16,25]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def get_subf_numerical_value(s):
    if s.upper().startswith('SUBF '):
        try: return int(s[5:])
        except: return float('inf')
    return float('inf')

# ─── Streamlit UI ─────────────────────────────────────────────────────────────
import streamlit.components.v1 as components

st.set_page_config(page_title="MOA Extractor — Hsenid", page_icon="📋", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
*{box-sizing:border-box;margin:0;padding:0;}
html,body,[class*="css"],.stApp{
    font-family:'Inter',sans-serif!important;
    background:#06070f!important;
    color:#f1f5f9!important;
}
/* Center the whole app like a square card */
.block-container{
    max-width:800px!important;
    padding:28px 0 48px!important;
    margin:0 auto!important;
}
header[data-testid="stHeader"]{display:none!important;}
.stDeployButton,footer,div[data-testid="stToolbar"]{display:none!important;}

/* Outer card */
.app-card{
    background:#0d0e1c;border:1px solid #1a1b30;
    border-radius:20px;padding:22px;display:flex;
    flex-direction:column;gap:14px;
}

/* TOPBAR */
.topbar{background:#13142a;border:1px solid #1e2040;border-radius:12px;padding:11px 16px;display:flex;align-items:center;justify-content:space-between;}
.tl{display:flex;align-items:center;gap:10px;}
.logo-box{width:32px;height:32px;border-radius:9px;background:linear-gradient(135deg,#6c63ff,#9c5cff);display:flex;align-items:center;justify-content:center;font-size:14px;font-weight:800;color:#fff;flex-shrink:0;}
.logo-name{font-size:14px;font-weight:800;color:#fff;}
.logo-name em{color:#7c6fff;font-style:normal;}
.logo-sub{font-size:9px;color:#3d4060;margin-top:1px;}
.tr{display:flex;align-items:center;gap:12px;}
.live-dot{width:8px;height:8px;background:#22c55e;border-radius:50%;animation:glow 2s infinite;flex-shrink:0;}
@keyframes glow{0%,100%{box-shadow:0 0 0 0 rgba(34,197,94,0.5)}50%{box-shadow:0 0 0 5px rgba(34,197,94,0)}}
.tr-right{text-align:right;}
.tr-time{font-size:15px;font-weight:700;color:#fff;font-variant-numeric:tabular-nums;line-height:1;}
.tr-date{font-size:10px;color:#3d4060;margin-top:2px;}
.tr-brand{font-size:11px;font-weight:700;color:#6c63ff;margin-top:2px;}

/* WELCOME ROW */
.welcome-row{display:flex;align-items:flex-start;justify-content:space-between;gap:14px;}
.wl{flex:1;min-width:0;}
.wl h2{font-size:20px;font-weight:800;color:#fff;margin-bottom:4px;}
.wl p{font-size:12px;color:#6b7280;margin-bottom:12px;}
.feats{display:flex;gap:14px;flex-wrap:wrap;}
.feat{display:flex;align-items:center;gap:8px;}
.fi{width:30px;height:30px;border-radius:8px;background:rgba(108,99,255,0.12);border:1px solid rgba(108,99,255,0.2);display:flex;align-items:center;justify-content:center;font-size:15px;flex-shrink:0;}
.ft{font-size:12px;font-weight:700;color:#e2e8f0;}
.fs{font-size:10px;color:#4b5563;}
.clocks{display:flex;gap:8px;flex-shrink:0;}
.cc{background:#13142a;border:1px solid #1e2040;border-radius:10px;padding:10px 14px;min-width:130px;}
.cc-top{display:flex;align-items:center;gap:6px;margin-bottom:5px;}
.cc-z{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;}
.cc-t{font-size:17px;font-weight:700;color:#fff;font-variant-numeric:tabular-nums;line-height:1;}
.cc-d{font-size:10px;color:#3d4060;margin-top:4px;}
.ist .cc-z{color:#6c63ff;}.fjt .cc-z{color:#8b5cf6;}.est .cc-z{color:#f59e0b;}

/* UPLOAD */
.upload-card{background:#13142a;border:1px solid #1e2040;border-radius:14px;padding:16px 16px 14px;}
.uc-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;}
.uc-title{font-size:14px;font-weight:700;color:#fff;}
.uc-badge{font-size:10px;color:#4b5563;background:#0d0e1c;border:1px solid #1e2040;border-radius:5px;padding:3px 9px;}

/* Drop zone — this is the REAL Streamlit uploader, styled */
[data-testid="stFileUploadDropzone"]{
    background:#0d0e1c!important;
    border:2px dashed #252640!important;
    border-radius:12px!important;
    padding:32px 20px!important;
}
[data-testid="stFileUploadDropzone"]:hover{
    border-color:#6c63ff!important;
    background:rgba(108,99,255,0.03)!important;
}
[data-testid="stFileUploadDropzone"] svg{display:none!important;}
[data-testid="stFileUploadDropzone"] small{display:none!important;}
[data-testid="stFileUploaderDropzoneInstructions"]{display:none!important;}
/* Show our custom content above the native uploader */
.drop-overlay{
    border:2px dashed #252640;border-radius:12px;
    padding:34px 20px 26px;text-align:center;background:#0d0e1c;
    position:relative;cursor:pointer;margin-bottom:0;
}
.drop-overlay:hover{border-color:#6c63ff;background:rgba(108,99,255,0.03);}
.drop-t1{font-size:17px;font-weight:700;color:#fff;margin-bottom:4px;}
.drop-t2{font-size:12px;color:#6b7280;margin-bottom:3px;}
.drop-t3{font-size:11px;color:#3d4060;margin-bottom:18px;}

/* Style the native Streamlit file uploader button as "Choose PDF File" */
.stFileUploader label{display:none!important;}
.stFileUploader section{background:transparent!important;}
[data-testid="baseButton-secondary"]{
    background:linear-gradient(135deg,#6c63ff,#8b5cf6)!important;
    color:#fff!important;border:none!important;
    border-radius:9px!important;padding:11px 26px!important;
    font-size:13px!important;font-weight:700!important;
    width:auto!important;
}

/* SUBF */
.subf-row{background:#13142a;border:1px solid #1e2040;border-radius:11px;padding:11px 14px;display:flex;align-items:center;justify-content:space-between;cursor:pointer;}
.sl{display:flex;align-items:center;gap:10px;}
.si{width:30px;height:30px;border-radius:7px;background:rgba(108,99,255,0.1);border:1px solid rgba(108,99,255,0.2);display:flex;align-items:center;justify-content:center;font-size:14px;flex-shrink:0;}
.st1{font-size:12px;font-weight:600;color:#e2e8f0;}
.st2{font-size:10px;color:#6c63ff;margin-top:1px;}

/* Extract btn */
.stButton>button{
    width:100%!important;padding:14px!important;border:none!important;
    border-radius:11px!important;font-size:14px!important;font-weight:700!important;
    color:#fff!important;background:linear-gradient(135deg,#6c63ff,#8b5cf6)!important;
    letter-spacing:0.02em!important;
}
.stButton>button:disabled{background:#1a1b30!important;color:#374151!important;}
.stButton>button:hover{filter:brightness(1.1)!important;}

/* File selected */
.file-selected{
    background:rgba(108,99,255,0.08);border:1px solid rgba(108,99,255,0.2);
    border-radius:9px;padding:9px 12px;margin-top:8px;
    font-size:11px;color:#a78bfa;
}

/* Security */
.sec-bar{background:#13142a;border:1px solid #1e2040;border-radius:12px;padding:13px 16px;display:flex;align-items:center;justify-content:space-between;}
.sec-l{display:flex;align-items:center;gap:10px;}
.sec-icon{width:32px;height:32px;border-radius:8px;background:rgba(34,197,94,0.08);border:1px solid rgba(34,197,94,0.15);display:flex;align-items:center;justify-content:center;font-size:17px;flex-shrink:0;}
.sec-t1{font-size:12px;font-weight:600;color:#e2e8f0;}
.sec-t2{font-size:10px;color:#3d4060;margin-top:2px;}

/* Progress */
.stProgress>div>div{background:linear-gradient(90deg,#6c63ff,#8b5cf6)!important;}

/* Result */
.result-box{background:rgba(108,99,255,0.07);border:1px solid rgba(108,99,255,0.25);border-radius:12px;padding:16px;margin-bottom:12px;}
.rb-t{font-size:13px;font-weight:700;color:#a78bfa;margin-bottom:10px;text-align:center;}
.rb-s{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;}
.rbs{background:rgba(0,0,0,0.3);border-radius:7px;padding:9px;text-align:center;}
.rbs-v{font-size:16px;font-weight:700;color:#fff;display:block;}
.rbs-l{font-size:9px;color:#4b5563;text-transform:uppercase;letter-spacing:0.06em;}

/* Download */
.stDownloadButton>button{
    width:100%!important;padding:13px!important;border-radius:10px!important;
    font-size:13px!important;font-weight:700!important;color:#fff!important;
    background:linear-gradient(135deg,#059669,#10b981)!important;border:none!important;
}

/* Expander */
.streamlit-expanderHeader{background:#0d0e1c!important;border:1px solid #1e2040!important;border-radius:8px!important;color:#e2e8f0!important;}
.stTextArea textarea{background:#0d0e1c!important;border:1px solid #1e2040!important;color:#f1f5f9!important;border-radius:8px!important;font-size:12px!important;}
</style>
""", unsafe_allow_html=True)

# ── START CARD ──
st.markdown('<div class="app-card">', unsafe_allow_html=True)

# ── TOPBAR ──
st.markdown("""
<div class="topbar">
  <div class="tl">
    <div class="logo-box">M</div>
    <div>
      <div class="logo-name">MOA <em>Extractor</em></div>
      <div class="logo-sub">PO Count Sheet Extractor</div>
    </div>
  </div>
  <div class="tr">
    <div class="live-dot"></div>
    <div class="tr-right">
      <div class="tr-time" id="nav-t">--:--:-- --</div>
      <div class="tr-date" id="nav-d">--</div>
      <div class="tr-brand">Hsenid</div>
    </div>
    <span style="font-size:18px">☀️</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── WELCOME + CLOCKS ──
st.markdown("""
<div class="welcome-row">
  <div class="wl">
    <h2>Welcome back! 👋</h2>
    <p>Upload your PO PDF and get a formatted Excel file instantly.</p>
    <div class="feats">
      <div class="feat"><div class="fi">🔍</div><div><div class="ft">No steps</div><div class="fs">Just upload your file</div></div></div>
      <div class="feat"><div class="fi">💻</div><div><div class="ft">No code</div><div class="fs">We handle the rest</div></div></div>
      <div class="feat"><div class="fi">⚡</div><div><div class="ft">Instant Excel</div><div class="fs">Download in seconds</div></div></div>
    </div>
  </div>
  <div class="clocks">
    <div class="cc ist">
      <div class="cc-top"><span style="font-size:13px">🇮🇳</span><span class="cc-z">IST</span></div>
      <div class="cc-t" id="ist-t">--:--:--</div>
      <div class="cc-d" id="ist-d">--</div>
    </div>
    <div class="cc fjt">
      <div class="cc-top"><span style="font-size:13px">🇫🇯</span><span class="cc-z">FJT</span></div>
      <div class="cc-t" id="fjt-t">--:--:--</div>
      <div class="cc-d" id="fjt-d">--</div>
    </div>
    <div class="cc est">
      <div class="cc-top"><span style="font-size:13px">🇺🇸</span><span class="cc-z">EST</span></div>
      <div class="cc-t" id="est-t">--:--:--</div>
      <div class="cc-d" id="est-d">--</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── UPLOAD CARD ──
st.markdown("""
<div class="upload-card">
  <div class="uc-head">
    <div class="uc-title">Upload PO PDF</div>
    <div class="uc-badge">Max 200MB per file</div>
  </div>
  <div class="drop-overlay">
    <svg width="58" height="44" viewBox="0 0 58 44" fill="none" style="display:block;margin:0 auto 12px" xmlns="http://www.w3.org/2000/svg">
      <path d="M44 36H14C7.925 36 3 31.075 3 25C3 19.17 7.5 14.35 13.22 13.78C14.56 8.5 19.33 4.5 25 4.5C32.456 4.5 38.5 10.544 38.5 18C38.5 18.25 38.49 18.5 38.47 18.75C42.13 19.19 45 22.27 45 26C45 26 47 26 47 28C47 32.418 44.418 36 44 36Z" fill="rgba(108,99,255,0.15)" stroke="#6c63ff" stroke-width="1.8"/>
      <path d="M29 28V19M29 19L24 24M29 19L34 24" stroke="#6c63ff" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"/>
    </svg>
    <div class="drop-t1">Drag &amp; drop your PDF here</div>
    <div class="drop-t2">Or click to browse files</div>
    <div class="drop-t3">Only PDF files 📋 Up to 200MB per file</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── REAL FILE UPLOADER (functional drag & drop + click) ──
uploaded_files = st.file_uploader(
    "Upload PDF files",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)

if uploaded_files:
    names = ", ".join(f.name for f in uploaded_files)
    st.markdown(f"""
    <div class="file-selected">
      ✅ &nbsp;<strong style="color:#fff">{len(uploaded_files)} file(s) ready:</strong> {names}
    </div>
    """, unsafe_allow_html=True)

# ── SUBF ROW ──
st.markdown("""
<div class="subf-row">
  <div class="sl">
    <div class="si">#️⃣</div>
    <div><div class="st1">Optional: SUBF number mapping</div><div class="st2">Improve extraction accuracy</div></div>
  </div>
  <span style="color:#374151;font-size:17px">›</span>
</div>
""", unsafe_allow_html=True)

subf_text = ""
with st.expander("SUBF Number Mapping", expanded=False):
    subf_text = st.text_area(
        "s", placeholder="SUBF 103566\t88251-24\nSUBF 103567\t88251-25",
        height=75, label_visibility="collapsed"
    )

# ── EXTRACT BUTTON ──
extract_btn = st.button("⚡  Extract to Excel", disabled=not uploaded_files, use_container_width=True)

# ── SECURITY BAR ──
st.markdown("""
<div class="sec-bar">
  <div class="sec-l">
    <div class="sec-icon">🛡️</div>
    <div>
      <div class="sec-t1">Your data is secure</div>
      <div class="sec-t2">We never store your files. All processing is done in real-time and files are automatically cleared.</div>
    </div>
  </div>
  <span style="font-size:24px">✅</span>
</div>
""", unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)  # close app-card

# ── LIVE CLOCKS via browser JS ──
components.html("""
<script>
(function(){
  function fT(d){return d.toLocaleTimeString('en-US',{hour:'2-digit',minute:'2-digit',second:'2-digit',hour12:true});}
  function fD(d){return d.toLocaleDateString('en-GB',{day:'2-digit',month:'short',year:'numeric'});}
  function set(id,v){var e=window.parent.document.getElementById(id);if(e)e.textContent=v;}
  function tick(){
    var now=new Date();
    var ist=new Date(now.toLocaleString('en-US',{timeZone:'Asia/Kolkata'}));
    var fjt=new Date(now.toLocaleString('en-US',{timeZone:'Pacific/Fiji'}));
    var est=new Date(now.toLocaleString('en-US',{timeZone:'America/New_York'}));
    set('nav-t',fT(now));set('nav-d',fD(now));
    set('ist-t',fT(ist));set('ist-d',fD(ist));
    set('fjt-t',fT(fjt));set('fjt-d',fD(fjt));
    set('est-t',fT(est));set('est-d',fD(est));
  }
  tick();setInterval(tick,1000);
})();
</script>
""", height=0)

# ── EXTRACTION LOGIC ──
if extract_btn and uploaded_files:
    progress_bar = st.progress(0, text="Starting extraction...")
    all_rows_by_key = defaultdict(list)
    all_order_numbers = []
    total_files = len(uploaded_files)

    for fi, uf in enumerate(uploaded_files):
        pdf_bytes = uf.read()
        def progress_cb(done, total, fi=fi, fname=uf.name):
            overall = int(((fi + done/max(total,1)) / total_files) * 90)
            progress_bar.progress(overall, text=f"📄 {fname} — page {done}/{total}")
        order_no, rows = parse_po(pdf_bytes, progress_cb=progress_cb)
        if order_no: all_order_numbers.append(order_no)
        for row in rows:
            key = (order_no, row['line'])
            all_rows_by_key[key].append(row)

    progress_bar.progress(92, text="Building Excel...")

    subf_num_map = {}
    subf_filter_keys = set()
    if subf_text and subf_text.strip():
        for entry in subf_text.strip().split('\n'):
            entry = entry.strip()
            if '\t' in entry:
                subf_label, order_line = entry.split('\t', 1)
                subf_number = subf_label.strip()
                if not subf_number.upper().startswith('SUBF'):
                    subf_number = 'SUBF ' + subf_number
                if '-' in order_line:
                    order, line = order_line.split('-', 1)
                    key = (order.strip(), line.strip())
                    subf_num_map[key] = subf_number
                    subf_filter_keys.add(key)

    if subf_filter_keys:
        all_rows_by_key = {k: v for k, v in all_rows_by_key.items() if k in subf_filter_keys}

    if not all_rows_by_key:
        st.error("No data extracted. Please check your PDF format.")
    else:
        if subf_filter_keys:
            keys_ordered = sorted(all_rows_by_key, key=lambda k: get_subf_numerical_value(subf_num_map.get(k,'')))
        else:
            keys_ordered = sorted(all_rows_by_key, key=lambda k: (k[0], int(k[1])))

        final_rows = []
        for key in keys_ordered:
            for row in all_rows_by_key[key]:
                row['subf_number'] = subf_num_map.get(key,'')
                final_rows.append(row)

        roster_counter = defaultdict(int)
        for item in final_rows:
            parts = [item.get('subf_number') or 'N/A_SUBF', item['size']]
            if item['scope'] == 'BASE DESIGN': parts.append('B')
            elif item['scope'] == 'ROSTER':
                k = (item.get('subf_number','N/A_SUBF'), item['size'])
                roster_counter[k] += 1
                parts.append(f'R{roster_counter[k]}')
            item['vis_name'] = '_'.join(parts)

        unique_orders = list(dict.fromkeys(all_order_numbers))
        if len(unique_orders)==1: out_name=unique_orders[0]
        elif len(unique_orders)>1: out_name=f'{unique_orders[0]}_and_{len(unique_orders)-1}_others'
        else: out_name='Combined'

        excel_bytes = build_excel(out_name, final_rows)
        output_filename = f'{out_name.replace(" ","_")}_count_sheet.xlsx'
        line_count = len(set(r['label'] for r in final_rows))
        roster_count = sum(r.get('qty',1) for r in final_rows if r['scope']=='ROSTER')
        progress_bar.progress(100, text="✅ Done!")

        st.markdown(f"""
        <div class="result-box">
          <div class="rb-t">✅ Excel Ready — {output_filename}</div>
          <div class="rb-s">
            <div class="rbs"><span class="rbs-v">{out_name}</span><span class="rbs-l">Order</span></div>
            <div class="rbs"><span class="rbs-v">{line_count}</span><span class="rbs-l">Lines</span></div>
            <div class="rbs"><span class="rbs-v">{len(final_rows)}</span><span class="rbs-l">Rows</span></div>
            <div class="rbs"><span class="rbs-v">{roster_count}</span><span class="rbs-l">Roster</span></div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.download_button(
            label="⬇️  Download Excel",
            data=excel_bytes,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
